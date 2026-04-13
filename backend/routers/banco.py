from fastapi import APIRouter, HTTPException, Depends, Query, File, UploadFile
from typing import List, Optional
from datetime import date
from database import get_pool
from models import Conciliacion, ConciliacionCreate
from dependencies import get_empresa_id, safe_date_param
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.post("/conciliacion/previsualizar-excel")
async def previsualizar_excel_banco(
    file: UploadFile = File(...),
    banco: str = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    import io
    from datetime import datetime as dt
    try:
        content = await file.read()
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        header_row = 1
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(row):
                row_str = ' '.join([str(c or '') for c in row]).lower()
                if 'fecha' in row_str or 'f. valor' in row_str or 'f. operacion' in row_str:
                    header_row = idx; break
        preview_data = []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 51, values_only=True):
            if not row or not any(row): continue
            fecha = descripcion = referencia = monto = None
            try:
                if banco == 'BCP':
                    fecha = row[1] if len(row) > 1 else None
                    descripcion = row[3] if len(row) > 3 else None
                    monto_val = row[4] if len(row) > 4 else None
                    referencia = row[7] if len(row) > 7 else None
                    if monto_val:
                        monto = float(monto_val) if not isinstance(monto_val, str) else float(str(monto_val).replace(',', ''))
                elif banco == 'BBVA':
                    concepto = row[5] if len(row) > 5 else None
                    if concepto and 'saldo final' in str(concepto).lower(): continue
                    fecha = row[1] if len(row) > 1 else None
                    referencia = row[4] if len(row) > 4 else None
                    descripcion = row[5] if len(row) > 5 else None
                    importe = row[6] if len(row) > 6 else None
                    if importe:
                        monto = float(importe) if not isinstance(importe, str) else float(str(importe).replace(',', ''))
                elif banco == 'IBK':
                    if len(row) < 10: continue
                    try:
                        nro = int(row[0]) if row[0] and str(row[0]).strip() else None
                        if not nro: continue
                    except (ValueError, TypeError): continue
                    fecha = row[1] if len(row) > 1 else None
                    referencia = row[3] if len(row) > 3 else None
                    descripcion = row[5] if len(row) > 5 else None
                    cargo = row[7] if len(row) > 7 else None
                    abono = row[8] if len(row) > 8 else None
                    if cargo and cargo != '' and str(cargo).strip() != 'nan':
                        try: monto = -abs(float(cargo) if not isinstance(cargo, str) else float(str(cargo).replace(',', '')))
                        except: monto = None
                    elif abono and abono != '' and str(abono).strip() != 'nan':
                        try: monto = abs(float(abono) if not isinstance(abono, str) else float(str(abono).replace(',', '')))
                        except: monto = None
                else:
                    fecha = row[0] if len(row) > 0 else None
                    descripcion = row[1] if len(row) > 1 else None
                    referencia = row[2] if len(row) > 2 else None
                    monto = row[3] if len(row) > 3 else None
                if fecha:
                    if isinstance(fecha, dt): fecha = fecha.date().isoformat()
                    elif hasattr(fecha, 'date'): fecha = fecha.date().isoformat()
                    elif isinstance(fecha, str):
                        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                            try: fecha = dt.strptime(fecha.strip(), fmt).date().isoformat(); break
                            except: continue
                if not fecha or monto is None: continue
                preview_data.append({"fecha": fecha, "banco": banco, "referencia": str(referencia)[:200] if referencia else "", "descripcion": str(descripcion)[:500] if descripcion else "", "monto": float(monto) if monto else 0.0})
                if len(preview_data) >= 50: break
            except Exception as row_error:
                logger.warning(f"Error parsing row: {row_error}"); continue
        return {"preview": preview_data, "total_rows": len(preview_data)}
    except Exception as e:
        logger.error(f"Error previewing Excel: {e}")
        raise HTTPException(500, f"Error al previsualizar: {str(e)}")


@router.post("/conciliacion/importar-excel")
async def importar_excel_banco(
    file: UploadFile = File(...),
    cuenta_financiera_id: int = Query(...),
    banco: str = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    import io
    from datetime import datetime as dt
    pool = await get_pool()
    try:
        content = await file.read()
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        async with pool.acquire() as conn:
            await conn.execute("SET search_path TO finanzas2, public")
            imported = updated = skipped = 0
            header_row = 1
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(row):
                    row_str = ' '.join([str(c or '') for c in row]).lower()
                    if 'fecha' in row_str or 'f. valor' in row_str or 'f. operacion' in row_str:
                        header_row = idx; break
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or not any(row): continue
                fecha = descripcion = referencia = monto = None
                try:
                    if banco == 'BCP':
                        fecha = row[1] if len(row) > 1 else None
                        descripcion = row[3] if len(row) > 3 else None
                        monto_val = row[4] if len(row) > 4 else None
                        referencia = row[7] if len(row) > 7 else None
                        if monto_val: monto = float(monto_val) if not isinstance(monto_val, str) else float(str(monto_val).replace(',', ''))
                    elif banco == 'BBVA':
                        concepto = row[5] if len(row) > 5 else None
                        if concepto and 'saldo final' in str(concepto).lower(): continue
                        fecha = row[1] if len(row) > 1 else None
                        referencia = row[4] if len(row) > 4 else None
                        descripcion = row[5] if len(row) > 5 else None
                        importe = row[6] if len(row) > 6 else None
                        if importe: monto = float(importe) if not isinstance(importe, str) else float(str(importe).replace(',', ''))
                    elif banco == 'IBK':
                        if len(row) < 10: continue
                        try:
                            nro = int(row[0]) if row[0] and str(row[0]).strip() else None
                            if not nro: continue
                        except: continue
                        fecha = row[1] if len(row) > 1 else None
                        referencia = row[3] if len(row) > 3 else None
                        descripcion = row[5] if len(row) > 5 else None
                        cargo = row[7] if len(row) > 7 else None
                        abono = row[8] if len(row) > 8 else None
                        if cargo and cargo != '' and str(cargo).strip() != 'nan':
                            try: monto = -abs(float(cargo) if not isinstance(cargo, str) else float(str(cargo).replace(',', '')))
                            except: monto = None
                        elif abono and abono != '' and str(abono).strip() != 'nan':
                            try: monto = abs(float(abono) if not isinstance(abono, str) else float(str(abono).replace(',', '')))
                            except: monto = None
                    else:
                        fecha = row[0] if len(row) > 0 else None
                        descripcion = row[1] if len(row) > 1 else None
                        referencia = row[2] if len(row) > 2 else None
                        monto = row[3] if len(row) > 3 else None
                    if fecha:
                        if isinstance(fecha, dt): fecha = fecha.date()
                        elif hasattr(fecha, 'date'): fecha = fecha.date()
                        elif isinstance(fecha, str):
                            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y']:
                                try: fecha = dt.strptime(fecha.strip(), fmt).date(); break
                                except: continue
                    if not fecha or monto is None: continue
                    ref_clean = str(referencia).strip()[:200] if referencia else ''
                    desc_clean = str(descripcion).strip()[:500] if descripcion else ''
                    existing = await conn.fetchrow("""
                        SELECT id, procesado FROM finanzas2.cont_banco_mov_raw
                        WHERE cuenta_financiera_id = $1 AND banco = $2 AND COALESCE(referencia, '') = $3 AND fecha = $4
                    """, cuenta_financiera_id, banco, ref_clean, fecha)
                    if existing:
                        if existing['procesado']: skipped += 1; continue
                        else:
                            await conn.execute("UPDATE finanzas2.cont_banco_mov_raw SET descripcion=$1, monto=$2, banco_excel=$3 WHERE id=$4", desc_clean, monto, banco, existing['id'])
                            updated += 1
                    else:
                        await conn.execute("""
                            INSERT INTO finanzas2.cont_banco_mov_raw
                            (cuenta_financiera_id, banco, fecha, descripcion, referencia, monto, banco_excel, procesado, empresa_id)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, FALSE, $8)
                        """, cuenta_financiera_id, banco, fecha, desc_clean, ref_clean if ref_clean else None, monto, banco, empresa_id)
                        imported += 1
                except Exception as row_error:
                    logger.warning(f"Error parsing row: {row_error}"); continue
        return {"message": f"Importados: {imported}, Actualizados: {updated}, Omitidos (ya conciliados): {skipped}", "imported": imported, "updated": updated, "skipped": skipped}
    except Exception as e:
        logger.error(f"Error importing Excel: {e}")
        raise HTTPException(500, f"Error al importar: {str(e)}")


@router.get("/conciliacion/historial")
async def get_historial_conciliaciones(empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        rows = await conn.fetch("""
            SELECT cl.id, cl.conciliacion_id, cl.banco_mov_id, cl.pago_id, cl.monto, cl.tipo, cl.conciliado, cl.created_at,
                   bm.fecha as fecha_banco, bm.descripcion as descripcion_banco, bm.referencia as ref_banco, bm.monto as monto_banco,
                   p.numero as numero_sistema, p.tipo as tipo_sistema, p.fecha as fecha_sistema, p.notas as descripcion_sistema, p.monto_total as monto_sistema,
                   cf.nombre as cuenta_nombre, cf.banco as banco
            FROM finanzas2.cont_conciliacion_linea cl
            LEFT JOIN finanzas2.cont_banco_mov_raw bm ON cl.banco_mov_id = bm.id
            LEFT JOIN finanzas2.cont_pago p ON cl.pago_id = p.id
            LEFT JOIN finanzas2.cont_conciliacion c ON cl.conciliacion_id = c.id
            LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
            WHERE cl.empresa_id = $1
            ORDER BY cl.created_at DESC
        """, empresa_id)
        result = []
        for r in rows:
            fecha_banco = r['fecha_banco']
            fecha_sistema = r['fecha_sistema']
            result.append({
                "id": r['id'], "conciliacion_id": r['conciliacion_id'], "banco_id": r['banco_mov_id'],
                "sistema_id": r['pago_id'], "banco_mov_id": r['banco_mov_id'], "pago_id": r['pago_id'],
                "fecha_banco": fecha_banco.isoformat() if fecha_banco else None,
                "fecha_sistema": fecha_sistema.isoformat() if fecha_sistema else None,
                "banco": r['banco'] or r['cuenta_nombre'] or '-',
                "ref_banco": r['ref_banco'] or '', "descripcion_banco": r['descripcion_banco'] or '',
                "monto": float(r['monto_banco'] or r['monto'] or 0),
                "numero_sistema": r['numero_sistema'] or '', "tipo_sistema": r['tipo_sistema'] or r['tipo'] or '',
                "descripcion_sistema": r['descripcion_sistema'] or '',
                "monto_sistema": float(r['monto_sistema'] or r['monto'] or 0),
                "conciliado": r['conciliado'],
            })
        return result


@router.post("/conciliacion/desconciliar")
async def desconciliar_movimientos(data: dict, empresa_id: int = Depends(get_empresa_id)):
    banco_id = data.get('banco_id')
    pago_id = data.get('pago_id')
    if not banco_id and not pago_id:
        raise HTTPException(400, "Se requiere al menos banco_id o pago_id")
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        if banco_id:
            await conn.execute("UPDATE finanzas2.cont_banco_mov_raw SET procesado=FALSE, conciliado=FALSE WHERE id=$1", banco_id)
        if pago_id:
            await conn.execute("UPDATE finanzas2.cont_pago SET conciliado=FALSE WHERE id=$1", pago_id)
        if banco_id and pago_id:
            await conn.execute("DELETE FROM finanzas2.cont_conciliacion_linea WHERE (banco_mov_id=$1 OR pago_id=$2) AND empresa_id=$3", banco_id, pago_id, empresa_id)
        elif pago_id:
            await conn.execute("DELETE FROM finanzas2.cont_conciliacion_linea WHERE pago_id=$1 AND empresa_id=$2", pago_id, empresa_id)
        elif banco_id:
            await conn.execute("DELETE FROM finanzas2.cont_conciliacion_linea WHERE banco_mov_id=$1 AND empresa_id=$2", banco_id, empresa_id)
    return {"message": "Movimientos desconciliados exitosamente"}


@router.get("/conciliacion/sugerencias")
async def sugerir_matches(
    cuenta_financiera_id: int = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    """Suggest automatic matches between pending bank and system movements (1:1, N:1, 1:N)."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")

        banco_movs = await conn.fetch("""
            SELECT id, fecha, referencia, descripcion, monto
            FROM finanzas2.cont_banco_mov_raw
            WHERE cuenta_financiera_id = $1 AND empresa_id = $2
              AND COALESCE(conciliado, false) = false
            ORDER BY fecha
        """, cuenta_financiera_id, empresa_id)

        sistema_movs = await conn.fetch("""
            SELECT id, fecha, referencia, notas, monto as monto_total, tipo
            FROM finanzas2.cont_movimiento_tesoreria
            WHERE cuenta_financiera_id = $1 AND empresa_id = $2
              AND COALESCE(conciliado, false) = false
            ORDER BY fecha
        """, cuenta_financiera_id, empresa_id)

        sugerencias = []
        banco_usado = set()
        sistema_usado = set()

        def monto_sistema_signed(sm):
            val = float(sm['monto_total'])
            return -val if sm['tipo'] == 'egreso' else val

        # Rule 1: Exact amount + exact reference (1:1)
        for bm in banco_movs:
            if bm['id'] in banco_usado:
                continue
            monto_banco = float(bm['monto'])
            ref_banco = (bm['referencia'] or '').strip().lower()
            if not ref_banco:
                continue
            for sm in sistema_movs:
                if sm['id'] in sistema_usado:
                    continue
                monto_sis = monto_sistema_signed(sm)
                ref_sis = (sm['referencia'] or '').strip().lower()
                if abs(monto_banco - monto_sis) < 0.01 and ref_sis and ref_banco == ref_sis:
                    sugerencias.append({
                        'banco_mov_ids': [bm['id']],
                        'sistema_mov_ids': [sm['id']],
                        'monto': monto_banco,
                        'regla': 'referencia_exacta',
                        'confianza': 'alta',
                        'tipo': '1:1'
                    })
                    banco_usado.add(bm['id'])
                    sistema_usado.add(sm['id'])
                    break

        # Rule 2: Exact amount + close date ±3 days (1:1)
        for bm in banco_movs:
            if bm['id'] in banco_usado:
                continue
            monto_banco = float(bm['monto'])
            fecha_banco = bm['fecha']
            for sm in sistema_movs:
                if sm['id'] in sistema_usado:
                    continue
                monto_sis = monto_sistema_signed(sm)
                fecha_sis = sm['fecha']
                if abs(monto_banco - monto_sis) < 0.01:
                    diff_days = abs((fecha_banco - fecha_sis).days)
                    if diff_days <= 3:
                        sugerencias.append({
                            'banco_mov_ids': [bm['id']],
                            'sistema_mov_ids': [sm['id']],
                            'monto': monto_banco,
                            'regla': 'monto_fecha',
                            'confianza': 'alta' if diff_days == 0 else 'media',
                            'tipo': '1:1'
                        })
                        banco_usado.add(bm['id'])
                        sistema_usado.add(sm['id'])
                        break

        # Rule 3: N:1 — Multiple sistema movements sum to one banco movement
        from itertools import combinations
        remaining_sistema = [sm for sm in sistema_movs if sm['id'] not in sistema_usado]
        for bm in banco_movs:
            if bm['id'] in banco_usado:
                continue
            monto_banco = float(bm['monto'])
            fecha_banco = bm['fecha']
            # Try combinations of 2 and 3 sistema movements
            found = False
            for combo_size in (2, 3, 4):
                if found or combo_size > len(remaining_sistema):
                    break
                for combo in combinations(remaining_sistema, combo_size):
                    if any(sm['id'] in sistema_usado for sm in combo):
                        continue
                    total_combo = sum(monto_sistema_signed(sm) for sm in combo)
                    if abs(monto_banco - total_combo) < 0.01:
                        # Check dates are within reasonable range (±7 days)
                        fechas_sis = [sm['fecha'] for sm in combo]
                        if all(abs((fecha_banco - f).days) <= 7 for f in fechas_sis):
                            sugerencias.append({
                                'banco_mov_ids': [bm['id']],
                                'sistema_mov_ids': [sm['id'] for sm in combo],
                                'monto': monto_banco,
                                'regla': f'{combo_size}_a_1',
                                'confianza': 'media',
                                'tipo': f'{combo_size}:1'
                            })
                            banco_usado.add(bm['id'])
                            for sm in combo:
                                sistema_usado.add(sm['id'])
                            remaining_sistema = [s for s in remaining_sistema if s['id'] not in sistema_usado]
                            found = True
                            break

        # Rule 4: 1:N — Multiple banco movements sum to one sistema movement
        remaining_banco = [bm for bm in banco_movs if bm['id'] not in banco_usado]
        for sm in sistema_movs:
            if sm['id'] in sistema_usado:
                continue
            monto_sis = monto_sistema_signed(sm)
            fecha_sis = sm['fecha']
            found = False
            for combo_size in (2, 3, 4):
                if found or combo_size > len(remaining_banco):
                    break
                for combo in combinations(remaining_banco, combo_size):
                    if any(bm['id'] in banco_usado for bm in combo):
                        continue
                    total_combo = sum(float(bm['monto']) for bm in combo)
                    if abs(monto_sis - total_combo) < 0.01:
                        fechas_bco = [bm['fecha'] for bm in combo]
                        if all(abs((fecha_sis - f).days) <= 7 for f in fechas_bco):
                            sugerencias.append({
                                'banco_mov_ids': [bm['id'] for bm in combo],
                                'sistema_mov_ids': [sm['id']],
                                'monto': monto_sis,
                                'regla': f'1_a_{combo_size}',
                                'confianza': 'media',
                                'tipo': f'1:{combo_size}'
                            })
                            sistema_usado.add(sm['id'])
                            for bm in combo:
                                banco_usado.add(bm['id'])
                            remaining_banco = [b for b in remaining_banco if b['id'] not in banco_usado]
                            found = True
                            break

        return {
            'sugerencias': sugerencias,
            'total': len(sugerencias),
            'pendientes_banco': len(banco_movs) - len(banco_usado),
            'pendientes_sistema': len(sistema_movs) - len(sistema_usado)
        }


@router.post("/conciliacion/confirmar-sugerencias")
async def confirmar_sugerencias(
    data: dict,
    empresa_id: int = Depends(get_empresa_id),
):
    """Confirm and persist a batch of suggested matches (supports 1:1, N:1, 1:N)."""
    sugerencias = data.get('sugerencias', [])
    if not sugerencias:
        raise HTTPException(400, "No hay sugerencias para confirmar")

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        async with conn.transaction():
            confirmados = 0
            for sug in sugerencias:
                # Support both old format (singular) and new format (arrays)
                banco_ids = sug.get('banco_mov_ids') or ([sug['banco_mov_id']] if sug.get('banco_mov_id') else [])
                sistema_ids = sug.get('sistema_mov_ids') or ([sug['sistema_mov_id']] if sug.get('sistema_mov_id') else [])
                if not banco_ids or not sistema_ids:
                    continue

                # Verify all are still pending
                all_valid = True
                cuenta_id = None
                total_monto = 0
                for bid in banco_ids:
                    bm = await conn.fetchrow(
                        "SELECT id, cuenta_financiera_id, monto FROM finanzas2.cont_banco_mov_raw WHERE id=$1 AND COALESCE(conciliado,false)=false",
                        bid)
                    if not bm:
                        all_valid = False; break
                    cuenta_id = bm['cuenta_financiera_id']
                    total_monto += float(bm['monto'])
                if not all_valid:
                    continue

                for sid in sistema_ids:
                    sm = await conn.fetchrow(
                        "SELECT id FROM finanzas2.cont_movimiento_tesoreria WHERE id=$1 AND COALESCE(conciliado,false)=false",
                        sid)
                    if not sm:
                        all_valid = False; break
                if not all_valid:
                    continue

                # Mark all as reconciled
                for bid in banco_ids:
                    await conn.execute(
                        "UPDATE finanzas2.cont_banco_mov_raw SET procesado=TRUE, conciliado=TRUE WHERE id=$1", bid)
                for sid in sistema_ids:
                    await conn.execute(
                        "UPDATE finanzas2.cont_movimiento_tesoreria SET conciliado=TRUE WHERE id=$1", sid)

                # Create single conciliacion record for this group
                from datetime import date as date_cls
                today = date_cls.today()
                tipo_match = sug.get('tipo', '1:1')
                conciliacion = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_conciliacion
                    (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_final, estado, notas, empresa_id)
                    VALUES ($1, $2, $2, $3, 'completado', $4, $5) RETURNING id
                """, cuenta_id, today, total_monto,
                    f"Auto-match ({tipo_match}): {sug.get('regla', 'auto')}", empresa_id)

                # Create lineas for each banco+sistema pair
                for bid in banco_ids:
                    bm_monto = await conn.fetchval("SELECT monto FROM finanzas2.cont_banco_mov_raw WHERE id=$1", bid)
                    for sid in sistema_ids:
                        await conn.execute("""
                            INSERT INTO finanzas2.cont_conciliacion_linea
                            (conciliacion_id, banco_mov_id, pago_id, tipo, monto, conciliado, empresa_id)
                            VALUES ($1, $2, $3, 'auto', $4, TRUE, $5)
                        """, conciliacion['id'], bid, sid, float(bm_monto or total_monto), empresa_id)

                confirmados += 1

    return {
        'message': f'{confirmados} sugerencias confirmadas exitosamente',
        'confirmados': confirmados
    }


@router.get("/conciliacion/movimientos-banco")
async def list_movimientos_banco(
    cuenta_financiera_id: Optional[int] = None,
    procesado: Optional[bool] = None,
    conciliado: Optional[bool] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        conditions = ["empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        if cuenta_financiera_id:
            conditions.append(f"cuenta_financiera_id = ${idx}"); params.append(cuenta_financiera_id); idx += 1
        if procesado is not None:
            conditions.append(f"procesado = ${idx}"); params.append(procesado); idx += 1
        if conciliado is not None:
            conditions.append(f"conciliado = ${idx}"); params.append(conciliado); idx += 1
        query = f"SELECT * FROM finanzas2.cont_banco_mov_raw WHERE {' AND '.join(conditions)} ORDER BY fecha DESC"
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]


@router.get("/conciliaciones", response_model=List[Conciliacion])
async def list_conciliaciones(cuenta_financiera_id: Optional[int] = None, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        if cuenta_financiera_id:
            rows = await conn.fetch("""
                SELECT c.*, cf.nombre as cuenta_nombre FROM finanzas2.cont_conciliacion c
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id
                WHERE c.cuenta_financiera_id = $1 ORDER BY c.fecha_fin DESC
            """, cuenta_financiera_id)
        else:
            rows = await conn.fetch("""
                SELECT c.*, cf.nombre as cuenta_nombre FROM finanzas2.cont_conciliacion c
                LEFT JOIN finanzas2.cont_cuenta_financiera cf ON c.cuenta_financiera_id = cf.id ORDER BY c.fecha_fin DESC
            """)
        result = []
        for row in rows:
            conc_dict = dict(row)
            lineas = await conn.fetch("SELECT * FROM finanzas2.cont_conciliacion_linea WHERE conciliacion_id = $1", row['id'])
            conc_dict['lineas'] = [dict(l) for l in lineas]
            result.append(conc_dict)
        return result


@router.post("/conciliacion/conciliar")
async def conciliar_movimientos(
    banco_ids: List[int] = Query(...),
    pago_ids: List[int] = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        async with conn.transaction():
            cuenta_id = None
            if banco_ids:
                mov = await conn.fetchrow("SELECT cuenta_financiera_id, fecha FROM finanzas2.cont_banco_mov_raw WHERE id=$1", banco_ids[0])
                if mov: cuenta_id = mov['cuenta_financiera_id']
            if banco_ids:
                await conn.execute("UPDATE finanzas2.cont_banco_mov_raw SET procesado=TRUE, conciliado=TRUE WHERE id=ANY($1::int[])", banco_ids)
            if pago_ids:
                try:
                    await conn.execute("ALTER TABLE finanzas2.cont_pago ADD COLUMN IF NOT EXISTS conciliado BOOLEAN DEFAULT FALSE")
                except: pass
                await conn.execute("UPDATE finanzas2.cont_pago SET conciliado=TRUE WHERE id=ANY($1::int[])", pago_ids)
            if cuenta_id:
                from datetime import date as date_cls
                today = date_cls.today()
                total_banco = 0
                if banco_ids:
                    result = await conn.fetchrow("SELECT SUM(monto) as total FROM finanzas2.cont_banco_mov_raw WHERE id=ANY($1::int[])", banco_ids)
                    total_banco = float(result['total']) if result['total'] else 0
                conciliacion = await conn.fetchrow("""
                    INSERT INTO finanzas2.cont_conciliacion
                    (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_final, estado, notas, empresa_id)
                    VALUES ($1, $2, $2, $3, 'completado', $4, $5) RETURNING id
                """, cuenta_id, today, total_banco, f"Conciliacion: {len(banco_ids)} mov. banco + {len(pago_ids)} mov. sistema", empresa_id)
                conciliacion_id = conciliacion['id']
                for i, pago_id in enumerate(pago_ids):
                    pago_info = await conn.fetchrow("SELECT monto_total FROM finanzas2.cont_pago WHERE id=$1", pago_id)
                    banco_mov_id = banco_ids[i] if i < len(banco_ids) else None
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_conciliacion_linea
                        (conciliacion_id, banco_mov_id, pago_id, tipo, monto, conciliado, empresa_id)
                        VALUES ($1, $2, $3, 'pago', $4, TRUE, $5)
                    """, conciliacion_id, banco_mov_id, pago_id, pago_info['monto_total'] if pago_info else 0, empresa_id)
                for i in range(len(pago_ids), len(banco_ids)):
                    banco_info = await conn.fetchrow("SELECT monto FROM finanzas2.cont_banco_mov_raw WHERE id=$1", banco_ids[i])
                    await conn.execute("""
                        INSERT INTO finanzas2.cont_conciliacion_linea
                        (conciliacion_id, banco_mov_id, tipo, monto, conciliado, empresa_id)
                        VALUES ($1, $2, 'banco', $3, TRUE, $4)
                    """, conciliacion_id, banco_ids[i], banco_info['monto'] if banco_info else 0, empresa_id)
        return {"message": f"Conciliados {len(banco_ids)} movimientos del banco y {len(pago_ids)} del sistema", "banco_conciliados": len(banco_ids), "sistema_conciliados": len(pago_ids)}


@router.post("/conciliacion/crear-gasto-bancario")
async def crear_gasto_desde_movimientos_bancarios(
    banco_ids: List[int] = Query(...),
    categoria_id: int = Query(...),
    descripcion: Optional[str] = Query("Gastos bancarios agrupados"),
    cuenta_financiera_id: int = Query(...),
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        async with conn.transaction():
            movimientos = await conn.fetch("SELECT * FROM finanzas2.cont_banco_mov_raw WHERE id=ANY($1::int[])", banco_ids)
            if not movimientos:
                raise HTTPException(404, "No se encontraron movimientos bancarios")
            already_conciliados = [m for m in movimientos if m['conciliado']]
            if already_conciliados:
                raise HTTPException(400, f"{len(already_conciliados)} movimientos ya estan conciliados")
            total = sum(abs(float(m['monto'])) for m in movimientos)
            last_gasto = await conn.fetchrow("SELECT numero FROM finanzas2.cont_gasto ORDER BY id DESC LIMIT 1")
            if last_gasto and last_gasto['numero']:
                try: numero = f"GAS-{int(last_gasto['numero'].split('-')[1]) + 1:06d}"
                except: numero = f"GAS-{len(movimientos):06d}"
            else: numero = "GAS-000001"
            gasto = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_gasto
                (numero, fecha, beneficiario_nombre, moneda_id, subtotal, igv, total, tipo_documento, numero_documento, notas, empresa_id)
                VALUES ($1, CURRENT_DATE, $2, 1, $3, 0, $3, 'gasto_bancario', $4, $5, $6) RETURNING id
            """, numero, 'Banco', total, numero, descripcion, empresa_id)
            gasto_id = gasto['id']
            await conn.execute("""
                INSERT INTO finanzas2.cont_gasto_linea (gasto_id, categoria_id, descripcion, importe, igv_aplica, empresa_id)
                VALUES ($1, $2, $3, $4, FALSE, $5)
            """, gasto_id, categoria_id, f"{descripcion} ({len(movimientos)} movimientos)", total, empresa_id)
            pago_numero = f"PAG-E-{numero}"
            pago = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_pago (numero, tipo, fecha, cuenta_financiera_id, moneda_id, monto_total, notas, empresa_id)
                VALUES ($1, 'egreso', CURRENT_DATE, $2, 1, $3, $4, $5) RETURNING id
            """, pago_numero, cuenta_financiera_id, total, f"Pago automatico de {descripcion}", empresa_id)
            pago_id = pago['id']
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_detalle (pago_id, cuenta_financiera_id, medio_pago, monto, empresa_id)
                VALUES ($1, $2, 'cargo_bancario', $3, $4)
            """, pago_id, cuenta_financiera_id, total, empresa_id)
            await conn.execute("""
                INSERT INTO finanzas2.cont_pago_aplicacion (pago_id, tipo_documento, documento_id, monto_aplicado, empresa_id)
                VALUES ($1, 'gasto', $2, $3, $4)
            """, pago_id, gasto_id, total, empresa_id)
            await conn.execute("UPDATE finanzas2.cont_banco_mov_raw SET procesado=TRUE, conciliado=TRUE WHERE id=ANY($1::int[])", banco_ids)
            try: await conn.execute("ALTER TABLE finanzas2.cont_pago ADD COLUMN IF NOT EXISTS conciliado BOOLEAN DEFAULT FALSE")
            except: pass
            await conn.execute("UPDATE finanzas2.cont_pago SET conciliado=TRUE WHERE id=$1", pago_id)
            from datetime import date as date_cls
            today = date_cls.today()
            conciliacion = await conn.fetchrow("""
                INSERT INTO finanzas2.cont_conciliacion (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_final, estado, notas, empresa_id)
                VALUES ($1, $2, $2, $3, 'completado', $4, $5) RETURNING id
            """, cuenta_financiera_id, today, total, f"Gasto bancario automatico: {descripcion} ({len(banco_ids)} movimientos)", empresa_id)
            conciliacion_id = conciliacion['id']
            await conn.execute("""
                INSERT INTO finanzas2.cont_conciliacion_linea (conciliacion_id, pago_id, tipo, documento_id, monto, conciliado, empresa_id)
                VALUES ($1, $2, 'gasto', $3, $4, TRUE, $5)
            """, conciliacion_id, pago_id, gasto_id, total, empresa_id)
        return {"message": f"Gasto creado exitosamente con {len(banco_ids)} movimientos bancarios", "gasto_id": gasto_id, "gasto_numero": numero, "pago_id": pago_id, "total": total, "movimientos_conciliados": len(banco_ids)}


@router.post("/conciliaciones", response_model=Conciliacion)
async def create_conciliacion(data: ConciliacionCreate, empresa_id: int = Depends(get_empresa_id)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        row = await conn.fetchrow("""
            INSERT INTO finanzas2.cont_conciliacion
            (cuenta_financiera_id, fecha_inicio, fecha_fin, saldo_inicial, saldo_final, notas, empresa_id)
            VALUES ($1, TO_DATE($2, 'YYYY-MM-DD'), TO_DATE($3, 'YYYY-MM-DD'), $4, $5, $6, $7) RETURNING *
        """, data.cuenta_financiera_id, safe_date_param(data.fecha_inicio), safe_date_param(data.fecha_fin),
            data.saldo_inicial, data.saldo_final, data.notas, empresa_id)
        result = dict(row)
        result['lineas'] = []
        return result
