from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import date
from database import get_pool
from dependencies import get_empresa_id
import io
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/export/compraapp")
async def export_compraapp(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    empresa_id: int = Depends(get_empresa_id),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        conditions = ["fp.empresa_id = $1"]
        params = [empresa_id]
        idx = 2
        if fecha_desde:
            conditions.append(f"fp.fecha_contable >= ${idx}")
            params.append(fecha_desde); idx += 1
        if fecha_hasta:
            conditions.append(f"fp.fecha_contable <= ${idx}")
            params.append(fecha_hasta); idx += 1
        facturas = await conn.fetch(f"""
            SELECT fp.*,
                   t.nombre as proveedor_nombre, t.numero_documento as proveedor_ruc,
                   t.tipo_documento as proveedor_tipo_doc,
                   m.codigo as moneda_codigo
            FROM finanzas2.cont_factura_proveedor fp
            LEFT JOIN finanzas2.cont_tercero t ON fp.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON fp.moneda_id = m.id
            WHERE {' AND '.join(conditions)}
            ORDER BY fp.fecha_contable ASC, fp.id ASC
        """, *params)
        gastos = await conn.fetch(f"""
            SELECT g.*,
                   t.nombre as proveedor_nombre, t.numero_documento as proveedor_ruc,
                   t.tipo_documento as proveedor_tipo_doc,
                   m.codigo as moneda_codigo
            FROM finanzas2.cont_gasto g
            LEFT JOIN finanzas2.cont_tercero t ON g.proveedor_id = t.id
            LEFT JOIN finanzas2.cont_moneda m ON g.moneda_id = m.id
            WHERE {' AND '.join(conditions).replace('fp.', 'g.')}
            ORDER BY g.fecha_contable ASC, g.id ASC
        """, *params)
        config = {}
        config_rows = await conn.fetch("SELECT clave, valor FROM finanzas2.cont_config WHERE empresa_id=$1", empresa_id)
        for r in config_rows:
            config[r['clave']] = r['valor']

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro de Compras"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "Periodo", "CUO", "Correlativo", "Fecha Emision", "Fecha Vcto", "Tipo Doc",
            "Serie", "Numero", "Tipo Doc Proveedor", "RUC Proveedor", "Razon Social",
            "Base Gravada", "IGV", "Base No Gravada", "ISC", "Otros", "Total",
            "Moneda", "Tipo Cambio", "Fecha Contable",
            "Tipo Retencion", "Tasa Retencion", "Base Retencion", "Monto Retencion"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        row_num = 2
        for fp in facturas:
            ret = await conn.fetchrow(
                "SELECT * FROM finanzas2.cont_retencion_detalle WHERE origen_tipo='factura' AND origen_id=$1 AND empresa_id=$2",
                fp['id'], empresa_id)

            fecha_emision = fp['fecha_factura']
            fecha_vcto = fp['fecha_vencimiento']
            fecha_contable = fp['fecha_contable'] or fp['fecha_factura']
            periodo = fecha_contable.strftime('%Y%m') if fecha_contable else ''
            tipo_doc = fp.get('tipo_comprobante_sunat') or '01'
            numero_raw = fp.get('numero') or ''
            serie = ''
            numero_val = numero_raw
            if '-' in numero_raw:
                parts = numero_raw.split('-', 1)
                serie = parts[0]
                numero_val = parts[1] if len(parts) > 1 else ''
            tipo_doc_prov = ''
            prov_tipo = fp.get('proveedor_tipo_doc', '')
            if prov_tipo == 'RUC': tipo_doc_prov = '6'
            elif prov_tipo == 'DNI': tipo_doc_prov = '1'
            elif prov_tipo == 'CE': tipo_doc_prov = '4'
            else: tipo_doc_prov = prov_tipo or '6'
            def fmt(v): return float(v) if v else 0.0
            data = [
                periodo, f"M{row_num-1}", str(row_num - 1),
                fecha_emision.strftime('%d/%m/%Y') if fecha_emision else '',
                fecha_vcto.strftime('%d/%m/%Y') if fecha_vcto else '',
                tipo_doc, serie, numero_val, tipo_doc_prov,
                fp.get('proveedor_ruc') or '', fp.get('proveedor_nombre') or '',
                fmt(fp.get('base_gravada')), fmt(fp.get('igv_sunat') or fp.get('igv')),
                fmt(fp.get('base_no_gravada')), fmt(fp.get('isc')), 0.0, fmt(fp.get('total')),
                fp.get('moneda_codigo') or 'PEN', fmt(fp.get('tipo_cambio')) or 1.0,
                fecha_contable.strftime('%d/%m/%Y') if fecha_contable else '',
                ret['codtasa'] if ret else '',
                float(ret['codtasa'] or 0) if ret and ret.get('codtasa') else '',
                float(ret['b_imp'] or 0) if ret and ret.get('b_imp') else '',
                float(ret['igv_ret'] or 0) if ret and ret.get('igv_ret') else '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
            row_num += 1

        for g in gastos:
            fecha_gasto = g.get('fecha')
            fecha_contable = g.get('fecha_contable') or fecha_gasto
            periodo = fecha_contable.strftime('%Y%m') if fecha_contable else ''
            tipo_doc = g.get('tipo_comprobante_sunat') or g.get('tipo_documento') or 'NV'
            numero_raw = g.get('numero_documento') or g.get('numero') or ''
            serie = ''
            numero_val = numero_raw
            if '-' in str(numero_raw):
                parts = str(numero_raw).split('-', 1)
                serie = parts[0]
                numero_val = parts[1] if len(parts) > 1 else ''
            tipo_doc_prov = ''
            prov_tipo = g.get('proveedor_tipo_doc', '')
            if prov_tipo == 'RUC': tipo_doc_prov = '6'
            elif prov_tipo == 'DNI': tipo_doc_prov = '1'
            elif prov_tipo == 'CE': tipo_doc_prov = '4'
            else: tipo_doc_prov = prov_tipo or ''
            def fmt_g(v): return float(v) if v else 0.0
            data = [
                periodo, f"M{row_num-1}", str(row_num - 1),
                fecha_gasto.strftime('%d/%m/%Y') if fecha_gasto else '',
                '',
                tipo_doc, serie, numero_val, tipo_doc_prov,
                g.get('proveedor_ruc') or '', g.get('beneficiario_nombre') or g.get('proveedor_nombre') or '',
                fmt_g(g.get('base_gravada', g.get('subtotal'))),
                fmt_g(g.get('igv_sunat') or g.get('igv')),
                fmt_g(g.get('base_no_gravada')), fmt_g(g.get('isc')), 0.0, fmt_g(g.get('total')),
                g.get('moneda_codigo') or 'PEN', fmt_g(g.get('tipo_cambio')) or 1.0,
                fecha_contable.strftime('%d/%m/%Y') if fecha_contable else '',
                '', '', '', '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
            row_num += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=registro_compras.xlsx"}
        )
