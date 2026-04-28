"""
Carga masiva por Excel para Gastos y Facturas de Proveedor.

Cada endpoint expone:
  GET  /<recurso>/template-excel   → descarga plantilla con cabeceras + ejemplo
  POST /<recurso>/import-excel      → recibe archivo, procesa filas, crea registros

Diseño:
  - Las columnas que referencian catálogos (Proveedor, Categoría, Línea, etc.)
    aceptan el NOMBRE del catálogo (case-insensitive). El backend lo resuelve a id.
  - Si el nombre no existe se reporta error en esa fila pero el resto continúa.
  - Devuelve un resumen: { creados, errores: [{fila, error}], creados_ids }
  - Cada importación usa una transacción por fila (no atómica global) — si una
    fila falla, las demás se siguen procesando.
"""
from io import BytesIO
from datetime import datetime, date, timedelta
from typing import Optional, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from database import get_pool
from dependencies import get_empresa_id, get_next_correlativo
from routers.gastos import _next_otro_correlativo
from routers.pagos import generate_pago_number

router = APIRouter()


# ─────────── Helpers ───────────

def _read_excel_rows(file_bytes: bytes) -> list[dict]:
    """Devuelve lista de dicts con columnas como llaves (limpias). Usa openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue  # fila vacía
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            if isinstance(v, str):
                v = v.strip()
                if not v:
                    v = None
            d[h] = v
        out.append(d)
    return out


def _to_float(v) -> float:
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in ('true', '1', 'si', 'sí', 'yes', 'y', 'x', 'aplica')


def _to_date(v) -> Optional[date]:
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


async def _resolve_proveedor(conn, empresa_id: int, nombre: Optional[str]) -> Optional[int]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        "SELECT id FROM finanzas2.cont_tercero WHERE LOWER(nombre)=LOWER($1) AND empresa_id=$2 LIMIT 1",
        nombre, empresa_id)
    return row['id'] if row else None


async def _resolve_categoria(conn, nombre: Optional[str], empresa_id: int) -> Optional[int]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        "SELECT id FROM finanzas2.cont_categoria WHERE LOWER(nombre)=LOWER($1) AND empresa_id=$2 LIMIT 1",
        nombre, empresa_id)
    return row['id'] if row else None


async def _resolve_categoria_gasto(conn, nombre: Optional[str], empresa_id: int) -> Optional[dict]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        "SELECT id, es_cif FROM finanzas2.cont_categoria_gasto WHERE LOWER(nombre)=LOWER($1) AND empresa_id=$2 LIMIT 1",
        nombre, empresa_id)
    return dict(row) if row else None


async def _resolve_linea_negocio(conn, nombre: Optional[str]) -> Optional[int]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        "SELECT id FROM finanzas2.cont_linea_negocio WHERE LOWER(nombre)=LOWER($1) LIMIT 1", nombre)
    return row['id'] if row else None


async def _resolve_centro_costo(conn, nombre: Optional[str], empresa_id: int) -> Optional[int]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        """SELECT id FROM finanzas2.cont_centro_costo
            WHERE (LOWER(nombre)=LOWER($1) OR LOWER(codigo)=LOWER($1))
              AND empresa_id=$2 LIMIT 1""", nombre, empresa_id)
    return row['id'] if row else None


async def _resolve_unidad_interna(conn, nombre: Optional[str], empresa_id: int) -> Optional[int]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        "SELECT id FROM finanzas2.fin_unidad_interna WHERE LOWER(nombre)=LOWER($1) AND empresa_id=$2 LIMIT 1",
        nombre, empresa_id)
    return row['id'] if row else None


async def _resolve_cuenta(conn, nombre: Optional[str], empresa_id: int) -> Optional[int]:
    if not nombre:
        return None
    row = await conn.fetchrow(
        "SELECT id FROM finanzas2.cont_cuenta_financiera WHERE LOWER(nombre)=LOWER($1) AND empresa_id=$2 LIMIT 1",
        nombre, empresa_id)
    return row['id'] if row else None


async def _resolve_moneda(conn, codigo: Optional[str]) -> Optional[int]:
    if not codigo:
        # Default PEN
        row = await conn.fetchrow("SELECT id FROM finanzas2.cont_moneda WHERE codigo='PEN' LIMIT 1")
        return row['id'] if row else None
    row = await conn.fetchrow(
        "SELECT id FROM finanzas2.cont_moneda WHERE LOWER(codigo)=LOWER($1) LIMIT 1", codigo)
    return row['id'] if row else None


def _build_xlsx(headers: list[str], example_row: list[Any], sheet_name: str = "Plantilla", instrucciones: list[str] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Header row
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 2)
    # Example row
    for col_idx, v in enumerate(example_row, start=1):
        ws.cell(row=2, column=col_idx, value=v)
    # Instrucciones
    if instrucciones:
        ws2 = wb.create_sheet("Instrucciones")
        ws2.column_dimensions['A'].width = 100
        ws2.cell(row=1, column=1, value="INSTRUCCIONES DE CARGA").font = Font(bold=True, size=14)
        for i, line in enumerate(instrucciones, start=3):
            ws2.cell(row=i, column=1, value=line)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────── GASTOS ───────────

GASTO_HEADERS = [
    "Fecha", "Tipo Doc", "N° Documento", "Proveedor", "Beneficiario",
    "Moneda", "Tipo Cambio", "Categoría Gasto", "Línea Categoría",
    "Descripción", "Importe", "IGV Aplica", "Línea Negocio",
    "Centro Costo", "Unidad Interna", "Cuenta Pago", "Medio Pago", "Notas",
]
GASTO_EJEMPLO = [
    "2026-04-27", "factura", "F001-00001", "Colortex SAC", "",
    "PEN", "1", "Transporte y combustible", "Transporte producción",
    "Pasaje a taller", 100.00, "TRUE", "Pantalon Denim - Element Premium",
    "CC-CORTE", "Corte Interno", "BCP Soles", "transferencia", "Pago de pasajes",
]
GASTO_INSTRUCCIONES = [
    "1. Llene una fila por cada gasto (NO modifique los nombres de las cabeceras).",
    "2. Fecha: formato YYYY-MM-DD (ej. 2026-04-27) o DD/MM/YYYY.",
    "3. Tipo Doc: uno de boleta, factura, recibo, ticket, otro.",
    "4. Si Tipo Doc='otro' y N° Documento está vacío, se autogenera correlativo (MM-YYYY-NNNN).",
    "5. Proveedor: NOMBRE exacto del proveedor en el sistema (case-insensitive).",
    "   Si está vacío, llene 'Beneficiario' (texto libre).",
    "6. Moneda: PEN o USD. Default PEN. Tipo Cambio se requiere si moneda=USD.",
    "7. Categoría Gasto: nombre de la categoría CIF/no-CIF (cabecera). Ej: 'Transporte y combustible', 'Marketing'.",
    "8. Línea Categoría: nombre de la categoría contable de la línea. Ej: 'Transporte producción'.",
    "9. Importe: con punto decimal. Ej: 100.50 (NO usar coma).",
    "10. IGV Aplica: TRUE / FALSE / SI / NO / 1 / 0.",
    "11. Línea Negocio, Centro Costo, Unidad Interna: nombres exactos del catálogo.",
    "12. Cuenta Pago: nombre de la cuenta financiera (BCP Soles, Caja Efectivo, etc.).",
    "13. Medio Pago: efectivo, transferencia, cheque, tarjeta. Default efectivo.",
    "14. Filas vacías se ignoran.",
    "",
    "El total del gasto se calcula automáticamente: Importe + IGV (si aplica).",
]


@router.get("/gastos/template-excel")
async def gasto_template_excel(empresa_id: int = Depends(get_empresa_id)):
    xlsx = _build_xlsx(GASTO_HEADERS, GASTO_EJEMPLO, "Gastos", GASTO_INSTRUCCIONES)
    return StreamingResponse(
        BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_gastos.xlsx"'},
    )


@router.post("/gastos/import-excel")
async def gasto_import_excel(
    file: UploadFile = File(...),
    empresa_id: int = Depends(get_empresa_id),
):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Solo archivos Excel (.xlsx)")
    contents = await file.read()
    try:
        rows = _read_excel_rows(contents)
    except Exception as e:
        raise HTTPException(400, f"Archivo inválido: {e}")

    if not rows:
        return {"creados": 0, "errores": [{"fila": 1, "error": "El archivo está vacío"}]}

    pool = await get_pool()
    creados = 0
    creados_ids: list[int] = []
    errores: list[dict] = []

    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        for idx, row in enumerate(rows, start=2):
            try:
                fecha = _to_date(row.get("Fecha"))
                if not fecha:
                    raise ValueError("Fecha inválida o vacía")
                tipo_doc = (row.get("Tipo Doc") or "boleta").strip().lower()
                if tipo_doc not in ("boleta", "factura", "recibo", "ticket", "otro"):
                    raise ValueError(f"Tipo Doc '{tipo_doc}' no reconocido")
                num_doc = (row.get("N° Documento") or "").strip() or None
                proveedor_id = await _resolve_proveedor(conn, empresa_id, row.get("Proveedor"))
                if row.get("Proveedor") and not proveedor_id:
                    raise ValueError(f"Proveedor '{row.get('Proveedor')}' no encontrado")
                beneficiario = row.get("Beneficiario") or None
                moneda_id = await _resolve_moneda(conn, row.get("Moneda"))
                tipo_cambio = _to_float(row.get("Tipo Cambio")) or None
                cat_gasto = await _resolve_categoria_gasto(conn, row.get("Categoría Gasto"), empresa_id)
                if row.get("Categoría Gasto") and not cat_gasto:
                    raise ValueError(f"Categoría Gasto '{row.get('Categoría Gasto')}' no encontrada")
                cat_linea = await _resolve_categoria(conn, row.get("Línea Categoría"), empresa_id)
                if row.get("Línea Categoría") and not cat_linea:
                    raise ValueError(f"Línea Categoría '{row.get('Línea Categoría')}' no encontrada")
                descripcion = row.get("Descripción") or None
                importe = _to_float(row.get("Importe"))
                if importe <= 0:
                    raise ValueError("Importe debe ser > 0")
                igv_aplica = _to_bool(row.get("IGV Aplica"))
                ln_id = await _resolve_linea_negocio(conn, row.get("Línea Negocio"))
                cc_id = await _resolve_centro_costo(conn, row.get("Centro Costo"), empresa_id)
                ui_id = await _resolve_unidad_interna(conn, row.get("Unidad Interna"), empresa_id)
                cuenta_id = await _resolve_cuenta(conn, row.get("Cuenta Pago"), empresa_id)
                medio_pago = (row.get("Medio Pago") or "efectivo").strip().lower()
                if medio_pago not in ("efectivo", "transferencia", "cheque", "tarjeta"):
                    medio_pago = "efectivo"
                notas = row.get("Notas") or None

                # Cálculos fiscales (Importes son SIN IGV en la plantilla)
                base_grav = importe if igv_aplica else 0.0
                base_no_grav = 0.0 if igv_aplica else importe
                igv = round(base_grav * 0.18, 2)
                subtotal = importe
                total = round(subtotal + igv, 2)

                # Auto-correlativo si tipo='otro' sin número
                if tipo_doc == 'otro' and not num_doc:
                    num_doc = await _next_otro_correlativo(conn, empresa_id, fecha)

                # Tipo de asignación
                tipo_asig = ('comun' if cat_gasto and cat_gasto.get('es_cif')
                             else ('directo' if ln_id else 'no_asignado'))

                async with conn.transaction():
                    # Generar número correlativo del gasto (GAS-YYYY-NNNNN)
                    year = fecha.year
                    numero = await get_next_correlativo(conn, empresa_id, 'gasto', f"GAS-{year}-")

                    grow = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_gasto
                        (empresa_id, numero, fecha, fecha_contable, beneficiario_nombre, proveedor_id,
                         moneda_id, subtotal, igv, total, tipo_documento, numero_documento, notas,
                         base_gravada, igv_sunat, base_no_gravada, isc, tipo_cambio,
                         categoria_gasto_id, tipo_asignacion, centro_costo_id, linea_negocio_id,
                         unidad_interna_id, cuenta_pago_id)
                        VALUES ($1,$2,TO_DATE($3,'YYYY-MM-DD'),TO_DATE($3,'YYYY-MM-DD'),$4,$5,
                                $6,$7,$8,$9,$10,$11,$12,
                                $13,$14,$15,$16,$17,
                                $18,$19,$20,$21,$22,$23)
                        RETURNING id
                    """, empresa_id, numero, fecha.isoformat(), beneficiario, proveedor_id,
                        moneda_id, subtotal, igv, total, tipo_doc, num_doc, notas,
                        base_grav, igv, base_no_grav, 0.0, tipo_cambio,
                        cat_gasto['id'] if cat_gasto else None, tipo_asig, cc_id, ln_id,
                        ui_id, cuenta_id)
                    gasto_id = grow['id']

                    await conn.execute("""
                        INSERT INTO finanzas2.cont_gasto_linea
                        (empresa_id, gasto_id, categoria_id, descripcion, importe, igv_aplica,
                         linea_negocio_id, centro_costo_id, unidad_interna_id)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                    """, empresa_id, gasto_id, cat_linea, descripcion, importe, igv_aplica,
                        ln_id, cc_id, ui_id)

                    # Si hay cuenta pago, crear pago + movimiento
                    if cuenta_id:
                        pago_numero = await generate_pago_number(conn, 'egreso', empresa_id)
                        prow = await conn.fetchrow("""
                            INSERT INTO finanzas2.cont_pago
                            (empresa_id, numero, tipo, fecha, cuenta_financiera_id, moneda_id,
                             monto_total, referencia, notas, centro_costo_id, linea_negocio_id)
                            VALUES ($1,$2,'egreso',TO_DATE($3,'YYYY-MM-DD'),$4,$5,$6,$7,$8,$9,$10)
                            RETURNING id
                        """, empresa_id, pago_numero, fecha.isoformat(), cuenta_id, moneda_id,
                            total, num_doc or numero, notas, cc_id, ln_id)
                        pago_id = prow['id']
                        await conn.execute("""
                            INSERT INTO finanzas2.cont_pago_detalle
                            (empresa_id, pago_id, cuenta_financiera_id, medio_pago, monto)
                            VALUES ($1,$2,$3,$4,$5)
                        """, empresa_id, pago_id, cuenta_id, medio_pago, total)
                        await conn.execute(
                            "UPDATE finanzas2.cont_cuenta_financiera SET saldo_actual = saldo_actual - $1 WHERE id = $2",
                            total, cuenta_id)

                    creados += 1
                    creados_ids.append(gasto_id)
            except Exception as e:
                errores.append({"fila": idx, "error": str(e)})

    return {"creados": creados, "errores": errores, "creados_ids": creados_ids,
            "total_filas": len(rows)}


# ─────────── FACTURAS PROVEEDOR ───────────

FACTURA_HEADERS = [
    "Fecha Factura", "Fecha Vencimiento", "Términos Días", "Tipo Doc", "N° Documento",
    "Proveedor", "Moneda", "Tipo Cambio", "Categoría", "Descripción",
    "Importe", "IGV Aplica", "Línea Negocio", "Centro Costo", "Unidad Interna",
    "Impuestos Incluidos", "Notas",
]
FACTURA_EJEMPLO = [
    "2026-04-27", "2026-05-27", 30, "factura", "F071-00009999",
    "Colortex SAC", "PEN", "1", "Transporte producción", "Compra de tela",
    1180.00, "TRUE", "Pantalon Denim - Element Premium", "CC-CORTE", "Corte Interno",
    "TRUE", "Factura inicial",
]
FACTURA_INSTRUCCIONES = [
    "1. Llene una fila por cada factura (NO modifique los nombres de las cabeceras).",
    "2. Fecha Factura: formato YYYY-MM-DD obligatorio.",
    "3. Fecha Vencimiento: opcional. Si está vacío se calcula con Términos Días.",
    "4. Tipo Doc: factura, boleta, recibo, nota_credito, nota_interna.",
    "5. N° Documento: si está vacío se autogenera (F-YYYY-NNNN).",
    "6. Proveedor: NOMBRE exacto del proveedor en el sistema.",
    "7. Categoría: nombre de la categoría contable de la línea.",
    "8. Importe: con punto decimal. Si 'Impuestos Incluidos'=TRUE, ya incluye IGV.",
    "9. IGV Aplica: TRUE si el ítem está afecto a IGV, FALSE si no.",
    "10. Impuestos Incluidos: TRUE si el importe ya incluye IGV (default), FALSE si es base gravada.",
    "11. Línea Negocio, Centro Costo, Unidad Interna: nombres exactos del catálogo.",
    "12. Filas vacías se ignoran.",
    "",
    "Total y CxP se calculan automáticamente.",
]


@router.get("/facturas-proveedor/template-excel")
async def factura_template_excel(empresa_id: int = Depends(get_empresa_id)):
    xlsx = _build_xlsx(FACTURA_HEADERS, FACTURA_EJEMPLO, "Facturas", FACTURA_INSTRUCCIONES)
    return StreamingResponse(
        BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_facturas.xlsx"'},
    )


@router.post("/facturas-proveedor/import-excel")
async def factura_import_excel(
    file: UploadFile = File(...),
    empresa_id: int = Depends(get_empresa_id),
):
    if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "Solo archivos Excel (.xlsx)")
    contents = await file.read()
    try:
        rows = _read_excel_rows(contents)
    except Exception as e:
        raise HTTPException(400, f"Archivo inválido: {e}")

    if not rows:
        return {"creados": 0, "errores": [{"fila": 1, "error": "El archivo está vacío"}]}

    pool = await get_pool()
    creados = 0
    creados_ids: list[int] = []
    errores: list[dict] = []

    async with pool.acquire() as conn:
        await conn.execute("SET search_path TO finanzas2, public")
        for idx, row in enumerate(rows, start=2):
            try:
                fecha = _to_date(row.get("Fecha Factura"))
                if not fecha:
                    raise ValueError("Fecha Factura inválida")
                fecha_venc = _to_date(row.get("Fecha Vencimiento"))
                terminos = int(row.get("Términos Días") or 30)
                if not fecha_venc:
                    fecha_venc = fecha + timedelta(days=terminos)
                tipo_doc = (row.get("Tipo Doc") or "factura").strip().lower()
                num_doc = (row.get("N° Documento") or "").strip() or None
                proveedor_id = await _resolve_proveedor(conn, empresa_id, row.get("Proveedor"))
                if row.get("Proveedor") and not proveedor_id:
                    raise ValueError(f"Proveedor '{row.get('Proveedor')}' no encontrado")
                if not proveedor_id:
                    raise ValueError("Proveedor requerido")
                moneda_id = await _resolve_moneda(conn, row.get("Moneda"))
                tipo_cambio = _to_float(row.get("Tipo Cambio")) or None
                categoria_id = await _resolve_categoria(conn, row.get("Categoría"), empresa_id)
                if row.get("Categoría") and not categoria_id:
                    raise ValueError(f"Categoría '{row.get('Categoría')}' no encontrada")
                descripcion = row.get("Descripción") or None
                importe = _to_float(row.get("Importe"))
                if importe <= 0:
                    raise ValueError("Importe debe ser > 0")
                igv_aplica = _to_bool(row.get("IGV Aplica"))
                impuestos_incluidos = _to_bool(row.get("Impuestos Incluidos"))
                ln_id = await _resolve_linea_negocio(conn, row.get("Línea Negocio"))
                cc_id = await _resolve_centro_costo(conn, row.get("Centro Costo"), empresa_id)
                ui_id = await _resolve_unidad_interna(conn, row.get("Unidad Interna"), empresa_id)
                notas = row.get("Notas") or None

                # Cálculo de totales
                if igv_aplica:
                    if impuestos_incluidos:
                        base_grav = round(importe / 1.18, 2)
                        igv = round(importe - base_grav, 2)
                        subtotal = base_grav
                        total = importe
                    else:
                        base_grav = importe
                        igv = round(importe * 0.18, 2)
                        subtotal = importe
                        total = round(importe + igv, 2)
                    base_no_grav = 0.0
                else:
                    base_grav = 0.0
                    igv = 0.0
                    base_no_grav = importe
                    subtotal = importe
                    total = importe

                # Número correlativo
                if not num_doc:
                    year = fecha.year
                    num_doc = await get_next_correlativo(conn, empresa_id, 'factura_proveedor', f"F-{year}-")

                async with conn.transaction():
                    # Verificar duplicado
                    dup = await conn.fetchval(
                        "SELECT 1 FROM finanzas2.cont_factura_proveedor WHERE numero=$1 AND empresa_id=$2",
                        num_doc, empresa_id)
                    if dup:
                        raise ValueError(f"Ya existe factura con número '{num_doc}'")

                    frow = await conn.fetchrow("""
                        INSERT INTO finanzas2.cont_factura_proveedor
                        (empresa_id, numero, proveedor_id, moneda_id, fecha_factura, fecha_contable, fecha_vencimiento,
                         terminos_dias, tipo_documento, estado, subtotal, igv, total, saldo_pendiente,
                         impuestos_incluidos, base_gravada, igv_sunat, base_no_gravada, isc, tipo_cambio,
                         notas, unidad_interna_id)
                        VALUES ($1,$2,$3,$4,TO_DATE($5,'YYYY-MM-DD'),TO_DATE($5,'YYYY-MM-DD'),TO_DATE($6,'YYYY-MM-DD'),
                                $7,$8,'pendiente',$9,$10,$11,$11,
                                $12,$13,$14,$15,$16,$17,$18,$19)
                        RETURNING id
                    """, empresa_id, num_doc, proveedor_id, moneda_id,
                        fecha.isoformat(), fecha_venc.isoformat(),
                        terminos, tipo_doc, subtotal, igv, total,
                        impuestos_incluidos, base_grav, igv, base_no_grav, 0.0, tipo_cambio,
                        notas, ui_id)
                    factura_id = frow['id']

                    await conn.execute("""
                        INSERT INTO finanzas2.cont_factura_proveedor_linea
                        (empresa_id, factura_id, categoria_id, descripcion, importe, igv_aplica,
                         linea_negocio_id, centro_costo_id, unidad_interna_id, tipo_linea)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,'categoria')
                    """, empresa_id, factura_id, categoria_id, descripcion, importe, igv_aplica,
                        ln_id, cc_id, ui_id)

                    await conn.execute("""
                        INSERT INTO finanzas2.cont_cxp
                        (empresa_id, factura_id, proveedor_id, monto_original, saldo_pendiente,
                         fecha_vencimiento, estado)
                        VALUES ($1,$2,$3,$4,$4,TO_DATE($5,'YYYY-MM-DD'),'pendiente')
                    """, empresa_id, factura_id, proveedor_id, total, fecha_venc.isoformat())

                    creados += 1
                    creados_ids.append(factura_id)
            except Exception as e:
                errores.append({"fila": idx, "error": str(e)})

    return {"creados": creados, "errores": errores, "creados_ids": creados_ids,
            "total_filas": len(rows)}
